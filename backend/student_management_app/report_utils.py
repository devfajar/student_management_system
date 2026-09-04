import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from student_management_app.models import StudentResult, AttendanceReport


def calculate_grade(total_marks):
    if total_marks >= 90:
        return 'A+', 4.0, 'Distinction'
    elif total_marks >= 80:
        return 'A', 3.75, 'Excellent'
    elif total_marks >= 70:
        return 'B', 3.0, 'Good'
    elif total_marks >= 60:
        return 'C', 2.0, 'Average'
    elif total_marks >= 50:
        return 'D', 1.0, 'Pass'
    else:
        return 'F', 0.0, 'Fail'

def generate_student_report_card_pdf(student):
    """
    Generates a PDF academic transcript report card for a student.
    Returns bytes of the generated PDF.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )

    story = []
    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle(
        'MainTitle',
        parent=styles['Heading1'],
        fontSize=20,
        leading=24,
        textColor=colors.HexColor('#1e3a8a'),
        alignment=1, # Centered
        fontName='Helvetica-Bold'
    )

    subtitle_style = ParagraphStyle(
        'SubTitle',
        parent=styles['Normal'],
        fontSize=11,
        leading=14,
        textColor=colors.HexColor('#475569'),
        alignment=1,
        fontName='Helvetica'
    )

    section_header = ParagraphStyle(
        'SectionHeader',
        parent=styles['Heading2'],
        fontSize=13,
        leading=16,
        textColor=colors.HexColor('#1e40af'),
        fontName='Helvetica-Bold',
        spaceBefore=10,
        spaceAfter=6
    )

    cell_bold = ParagraphStyle(
        'CellBold',
        parent=styles['Normal'],
        fontSize=9,
        leading=11,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor('#1e293b')
    )

    cell_normal = ParagraphStyle(
        'CellNormal',
        parent=styles['Normal'],
        fontSize=9,
        leading=11,
        fontName='Helvetica',
        textColor=colors.HexColor('#334155')
    )

    # 1. Header & Branding
    story.append(Paragraph("STUDENT MANAGEMENT SYSTEM", title_style))
    story.append(Paragraph("OFFICIAL ACADEMIC TRANSCRIPT & PERFORMANCE REPORT", subtitle_style))
    story.append(Spacer(1, 8))
    story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#2563eb'), spaceBefore=2, spaceAfter=12))

    # 2. Student & Institutional Info
    course_name = student.course_id.course_name if student.course_id else "N/A"
    session_str = f"{student.session_year_id.session_start_year} - {student.session_year_id.session_end_year}" if student.session_year_id else "N/A"
    full_name = f"{student.admin.first_name} {student.admin.last_name}".strip() or student.admin.username

    info_data = [
        [
            Paragraph("<b>Student Name:</b>", cell_bold), Paragraph(full_name, cell_normal),
            Paragraph("<b>Student ID:</b>", cell_bold), Paragraph(f"#{student.id}", cell_normal),
        ],
        [
            Paragraph("<b>Username / Reg No:</b>", cell_bold), Paragraph(student.admin.username, cell_normal),
            Paragraph("<b>Gender:</b>", cell_bold), Paragraph(student.gender or "N/A", cell_normal),
        ],
        [
            Paragraph("<b>Enrolled Course:</b>", cell_bold), Paragraph(course_name, cell_normal),
            Paragraph("<b>Academic Session:</b>", cell_bold), Paragraph(session_str, cell_normal),
        ],
        [
            Paragraph("<b>Email:</b>", cell_bold), Paragraph(student.admin.email, cell_normal),
            Paragraph("<b>Date of Issue:</b>", cell_bold), Paragraph(datetime.now().strftime("%B %d, %Y"), cell_normal),
        ],
    ]

    info_table = Table(info_data, colWidths=[120, 160, 110, 150])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8fafc')),
        ('PADDING', (0, 0), (-1, -1), 5),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 14))

    # 3. Attendance Overview
    att_reports = AttendanceReport.objects.filter(student_id=student)
    total_classes = att_reports.count()
    present_classes = att_reports.filter(status=True).count()
    attendance_pct = (present_classes / total_classes * 100) if total_classes > 0 else 0.0

    att_summary_data = [
        [
            Paragraph(f"<b>Classes Conducted:</b> {total_classes}", cell_normal),
            Paragraph(f"<b>Present Days:</b> {present_classes}", cell_normal),
            Paragraph(f"<b>Attendance Rate:</b> {attendance_pct:.1f}%", cell_bold)
        ]
    ]
    att_table = Table(att_summary_data, colWidths=[180, 180, 180])
    att_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#eff6ff')),
        ('PADDING', (0, 0), (-1, -1), 6),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#bfdbfe')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ]))
    story.append(att_table)
    story.append(Spacer(1, 12))

    # 4. Examination Results Table
    story.append(Paragraph("Coursework & Examination Performance", section_header))

    results = StudentResult.objects.filter(student_id=student).select_related('subject_id')
    marks_table_data = [
        [
            Paragraph("<b>Subject / Course Module</b>", cell_bold),
            Paragraph("<b>Exam (50)</b>", cell_bold),
            Paragraph("<b>Assign (50)</b>", cell_bold),
            Paragraph("<b>Total (100)</b>", cell_bold),
            Paragraph("<b>Grade</b>", cell_bold),
            Paragraph("<b>GPA</b>", cell_bold),
            Paragraph("<b>Status</b>", cell_bold)
        ]
    ]

    total_gpa_points = 0.0
    subject_count = 0

    if results.exists():
        for res in results:
            exam = res.subject_exam_marks or 0.0
            assign = res.subject_assignment_marks or 0.0
            total = exam + assign
            letter_grade, gpa_point, remark = calculate_grade(total)

            total_gpa_points += gpa_point
            subject_count += 1

            status_color = '#15803d' if total >= 50 else '#b91c1c'
            status_text = f"<font color='{status_color}'><b>{'PASS' if total >= 50 else 'FAIL'}</b></font>"

            marks_table_data.append([
                Paragraph(res.subject_id.subject_name if res.subject_id else "N/A", cell_normal),
                Paragraph(f"{exam:.1f}", cell_normal),
                Paragraph(f"{assign:.1f}", cell_normal),
                Paragraph(f"{total:.1f}", cell_bold),
                Paragraph(letter_grade, cell_bold),
                Paragraph(f"{gpa_point:.2f}", cell_normal),
                Paragraph(status_text, cell_normal)
            ])
    else:
        marks_table_data.append([
            Paragraph("<i>No examination results recorded for this period</i>", cell_normal),
            "", "", "", "", "", ""
        ])

    marks_table = Table(marks_table_data, colWidths=[190, 60, 60, 65, 50, 50, 65])
    marks_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e2e8f0')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
        ('PADDING', (0, 0), (-1, -1), 5),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(marks_table)
    story.append(Spacer(1, 14))

    # 5. CGPA & Academic Summary Card
    cgpa = (total_gpa_points / subject_count) if subject_count > 0 else 0.0
    cgpa_letter, _, cgpa_remark = calculate_grade(cgpa * 25) # Scale to 100 for remark

    summary_data = [
        [
            Paragraph(f"<b>Subjects Evaluated:</b> {subject_count}", cell_normal),
            Paragraph(f"<b>Cumulative GPA:</b> <font size='11' color='#1e3a8a'><b>{cgpa:.2f} / 4.00</b></font>", cell_normal),
            Paragraph(f"<b>Standing:</b> <b>{cgpa_remark} ({cgpa_letter})</b>", cell_normal)
        ]
    ]
    summary_table = Table(summary_data, colWidths=[180, 180, 180])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f1f5f9')),
        ('PADDING', (0, 0), (-1, -1), 8),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#94a3b8')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 30))

    # 6. Signatures & Verification Stamp
    sig_data = [
        [
            Paragraph("____________________________<br/><b>Registrar / Academic Dean</b>", cell_normal),
            Paragraph("____________________________<br/><b>Controller of Examinations</b>", cell_normal),
            Paragraph("<b>Official Digital Seal</b><br/><i>Verified System Document</i>", cell_normal)
        ]
    ]
    sig_table = Table(sig_data, colWidths=[180, 180, 180])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'BOTTOM'),
        ('PADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(sig_table)

    # Build PDF
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


# ==========================================
# Excel (.xlsx) Export Engine
# ==========================================

def _apply_excel_styling(ws, header_title, columns):
    """
    Applies standard branding header styling, thin borders, and auto column widths to an openpyxl worksheet.
    """
    header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Style header row 1
    for col_num in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[1].height = 26

    # Auto-adjust column width
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Style all data rows
    if ws.max_row >= 2:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(columns)):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')


def generate_attendance_excel_bytes(attendance_reports):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Log"

    columns = [
        "Record ID", "Student ID", "Username", "Full Name",
        "Subject", "Attendance Date", "Status", "Recorded At"
    ]
    ws.append(columns)

    for rep in attendance_reports:
        stud = rep.student_id
        admin_user = stud.admin if stud else None
        att = rep.attendance_id
        ws.append([
            rep.id,
            stud.id if stud else "N/A",
            admin_user.username if admin_user else "N/A",
            f"{admin_user.first_name} {admin_user.last_name}".strip() if admin_user else "N/A",
            att.subject_id.subject_name if att and att.subject_id else "N/A",
            att.attendance_date.strftime("%Y-%m-%d") if att and att.attendance_date else "N/A",
            "Present" if rep.status else "Absent",
            rep.created_at.strftime("%Y-%m-%d %H:%M") if rep.created_at else "N/A"
        ])

    _apply_excel_styling(ws, "Attendance Log Report", columns)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_fees_excel_bytes(invoices):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fee Invoices"

    columns = [
        "Invoice ID", "Student ID", "Username", "Student",
        "Course", "Fee Title", "Total Amount", "Paid Amount",
        "Balance", "Payment Status", "Due Date", "Created Date"
    ]
    ws.append(columns)

    for inv in invoices:
        stud = inv.student_id
        admin_user = stud.admin if stud else None
        fee = inv.fee_structure_id
        ws.append([
            inv.id,
            stud.id if stud else "N/A",
            admin_user.username if admin_user else "N/A",
            f"{admin_user.first_name} {admin_user.last_name}".strip() if admin_user else "N/A",
            stud.course_id.course_name if stud and stud.course_id else "N/A",
            fee.fee_name if fee else "N/A",
            float(inv.total_amount),
            float(inv.paid_amount),
            float(inv.balance_amount),
            inv.payment_status,
            fee.due_date.strftime("%Y-%m-%d") if fee and fee.due_date else "N/A",
            inv.created_at.strftime("%Y-%m-%d") if inv.created_at else "N/A"
        ])

    _apply_excel_styling(ws, "Fee Ledger Report", columns)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_students_excel_bytes(students):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students Roster"

    columns = [
        "Student ID", "Username", "Full Name", "Email",
        "Gender", "Course", "Academic Session", "Address", "Enrolled Date"
    ]
    ws.append(columns)

    for stud in students:
        admin_user = stud.admin
        sess = stud.session_year_id
        ws.append([
            stud.id,
            admin_user.username if admin_user else "N/A",
            f"{admin_user.first_name} {admin_user.last_name}".strip() if admin_user else "N/A",
            admin_user.email if admin_user else "N/A",
            stud.gender or "N/A",
            stud.course_id.course_name if stud.course_id else "N/A",
            f"{sess.session_start_year} - {sess.session_end_year}" if sess else "N/A",
            stud.address or "N/A",
            stud.created_at.strftime("%Y-%m-%d") if stud.created_at else "N/A"
        ])

    _apply_excel_styling(ws, "Student Roster Report", columns)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_results_excel_bytes(results):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Examination Results"

    columns = [
        "Result ID", "Student ID", "Username", "Student",
        "Course", "Subject", "Exam Marks", "Assignment Marks",
        "Total Score", "Grade", "Standing", "Status"
    ]
    ws.append(columns)

    for res in results:
        stud = res.student_id
        admin_user = stud.admin if stud else None
        total = float(res.subject_exam_marks or 0) + float(res.subject_assignment_marks or 0)
        grade, gpa, standing = calculate_grade(total)
        ws.append([
            res.id,
            stud.id if stud else "N/A",
            admin_user.username if admin_user else "N/A",
            f"{admin_user.first_name} {admin_user.last_name}".strip() if admin_user else "N/A",
            stud.course_id.course_name if stud and stud.course_id else "N/A",
            res.subject_id.subject_name if res.subject_id else "N/A",
            float(res.subject_exam_marks or 0),
            float(res.subject_assignment_marks or 0),
            total,
            grade,
            standing,
            "Pass" if total >= 50 else "Fail"
        ])

    _apply_excel_styling(ws, "Examination Results Report", columns)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ==========================================
# Staff Payslip & Payroll Export Engine
# ==========================================

import csv

def generate_payslip_pdf_bytes(payroll):
    """
    Generates a professional salary payslip PDF for an individual staff payroll record.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )

    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'MainTitle',
        parent=styles['Heading1'],
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#1e3a8a'),
        alignment=1,
        fontName='Helvetica-Bold'
    )
    subtitle_style = ParagraphStyle(
        'SubTitle',
        parent=styles['Normal'],
        fontSize=11,
        leading=15,
        textColor=colors.HexColor('#475569'),
        alignment=1
    )
    meta_style = ParagraphStyle(
        'MetaText',
        parent=styles['Normal'],
        fontSize=9,
        leading=13,
        textColor=colors.HexColor('#64748b'),
        alignment=1
    )
    cell_bold = ParagraphStyle(
        'CellBold',
        parent=styles['Normal'],
        fontSize=10,
        leading=13,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor('#1e293b')
    )
    cell_normal = ParagraphStyle(
        'CellNormal',
        parent=styles['Normal'],
        fontSize=9,
        leading=13,
        textColor=colors.HexColor('#334155')
    )
    money_bold = ParagraphStyle(
        'MoneyBold',
        parent=styles['Normal'],
        fontSize=10,
        leading=13,
        fontName='Helvetica-Bold',
        alignment=2, # Right
        textColor=colors.HexColor('#0f172a')
    )

    month_names = [
        "", "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December"
    ]
    month_str = month_names[payroll.payroll_month] if 1 <= payroll.payroll_month <= 12 else str(payroll.payroll_month)

    # Header Banner
    story.append(Paragraph("STUDENT MANAGEMENT SYSTEM", title_style))
    story.append(Spacer(1, 4))
    story.append(Paragraph(f"<b>OFFICIAL SALARY PAYSLIP &bull; {month_str.upper()} {payroll.payroll_year}</b>", subtitle_style))
    story.append(Spacer(1, 2))
    story.append(Paragraph(f"Issued On: {datetime.now().strftime('%B %d, %Y')} &bull; Document Reference: PAY-{payroll.id:05d}", meta_style))
    story.append(Spacer(1, 12))
    story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#1e3a8a'), spaceBefore=2, spaceAfter=14))

    # Employee Details Grid
    staff = payroll.staff
    admin_user = staff.admin if staff else None
    staff_name = f"{admin_user.first_name} {admin_user.last_name}".strip() if admin_user else "N/A"
    designation = getattr(getattr(staff, 'salary_structure', None), 'designation', 'Staff Faculty')

    status_color = '#15803d' if payroll.payment_status == 'Paid' else '#b45309'
    info_data = [
        [
            Paragraph(f"<b>Employee Name:</b> {staff_name}", cell_normal),
            Paragraph(f"<b>Employee ID:</b> EMP-{staff.id:04d}", cell_normal)
        ],
        [
            Paragraph(f"<b>Designation:</b> {designation}", cell_normal),
            Paragraph(f"<b>Email:</b> {admin_user.email if admin_user else 'N/A'}", cell_normal)
        ],
        [
            Paragraph(f"<b>Payroll Period:</b> {month_str} {payroll.payroll_year}", cell_normal),
            Paragraph(f"<b>Payment Status:</b> <font color='{status_color}'><b>{payroll.payment_status.upper()}</b></font>", cell_normal)
        ],
        [
            Paragraph(f"<b>Payment Method:</b> {payroll.payment_method}", cell_normal),
            Paragraph(f"<b>Payment Date:</b> {payroll.payment_date if payroll.payment_date else 'Pending'}", cell_normal)
        ]
    ]
    info_table = Table(info_data, colWidths=[270, 270])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8fafc')),
        ('PADDING', (0, 0), (-1, -1), 6),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e1')),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 16))

    # Earnings & Deductions Breakdown
    earnings_total = float(payroll.basic_salary or 0) + float(payroll.allowances or 0) + float(payroll.bonus or 0)
    deductions_total = float(payroll.deductions or 0)
    net_total = float(payroll.net_salary or 0)

    breakdown_data = [
        [
            Paragraph("<b>Earnings Category</b>", cell_bold),
            Paragraph("<b>Amount ($)</b>", money_bold),
            Paragraph("<b>Deductions Category</b>", cell_bold),
            Paragraph("<b>Amount ($)</b>", money_bold)
        ],
        [
            Paragraph("Basic Base Salary", cell_normal),
            Paragraph(f"${float(payroll.basic_salary or 0):,.2f}", money_bold),
            Paragraph("Statutory / Taxes / Deductions", cell_normal),
            Paragraph(f"${deductions_total:,.2f}", money_bold)
        ],
        [
            Paragraph("Allowances (Housing/Travel)", cell_normal),
            Paragraph(f"${float(payroll.allowances or 0):,.2f}", money_bold),
            Paragraph("", cell_normal),
            Paragraph("", money_bold)
        ],
        [
            Paragraph("Performance Bonus / Incentives", cell_normal),
            Paragraph(f"${float(payroll.bonus or 0):,.2f}", money_bold),
            Paragraph("", cell_normal),
            Paragraph("", money_bold)
        ],
        [
            Paragraph("<b>Total Gross Earnings</b>", cell_bold),
            Paragraph(f"<b>${earnings_total:,.2f}</b>", money_bold),
            Paragraph("<b>Total Deductions</b>", cell_bold),
            Paragraph(f"<b>${deductions_total:,.2f}</b>", money_bold)
        ]
    ]

    breakdown_table = Table(breakdown_data, colWidths=[170, 100, 170, 100])
    breakdown_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('PADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f1f5f9')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(breakdown_table)
    story.append(Spacer(1, 16))

    # Net Salary Highlight Box
    net_box_data = [
        [
            Paragraph("<b>NET TAKE-HOME PAYABLE:</b>", ParagraphStyle('NetLabel', parent=styles['Normal'], fontSize=12, fontName='Helvetica-Bold', textColor=colors.HexColor('#1e3a8a'))),
            Paragraph(f"<b>${net_total:,.2f} USD</b>", ParagraphStyle('NetVal', parent=styles['Normal'], fontSize=14, fontName='Helvetica-Bold', alignment=2, textColor=colors.HexColor('#15803d')))
        ]
    ]
    net_table = Table(net_box_data, colWidths=[340, 200])
    net_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#ecfdf5')),
        ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#10b981')),
        ('PADDING', (0, 0), (-1, -1), 10),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(net_table)
    story.append(Spacer(1, 14))

    if payroll.remarks:
        story.append(Paragraph(f"<b>Note:</b> {payroll.remarks}", cell_normal))
        story.append(Spacer(1, 14))

    story.append(Spacer(1, 24))

    # Signature Block
    sig_data = [
        [
            Paragraph("____________________________<br/><b>Finance & Payroll Officer</b>", cell_normal),
            Paragraph("____________________________<br/><b>Employee Signature</b>", cell_normal),
            Paragraph("<b>System Verified Stamp</b><br/><i>Auto-Generated Secure Payslip</i>", cell_normal)
        ]
    ]
    sig_table = Table(sig_data, colWidths=[180, 180, 180])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'BOTTOM'),
        ('PADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(sig_table)

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def generate_payroll_excel_bytes(payrolls, month=None, year=None):
    """
    Generates an Excel workbook for staff payroll records with styled headers.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Staff Payroll Ledger"

    columns = [
        "Payroll ID", "Staff ID", "Username", "Staff Name", "Designation",
        "Month/Year", "Basic Salary", "Allowances", "Bonus",
        "Deductions", "Net Salary", "Payment Status", "Payment Method", "Payment Date"
    ]
    ws.append(columns)

    for p in payrolls:
        staff = p.staff
        admin_user = staff.admin if staff else None
        designation = getattr(getattr(staff, 'salary_structure', None), 'designation', 'Staff')
        ws.append([
            p.id,
            staff.id if staff else "N/A",
            admin_user.username if admin_user else "N/A",
            f"{admin_user.first_name} {admin_user.last_name}".strip() if admin_user else "N/A",
            designation,
            f"{p.payroll_month}/{p.payroll_year}",
            float(p.basic_salary or 0),
            float(p.allowances or 0),
            float(p.bonus or 0),
            float(p.deductions or 0),
            float(p.net_salary or 0),
            p.payment_status,
            p.payment_method,
            str(p.payment_date) if p.payment_date else "N/A"
        ])

    _apply_excel_styling(ws, "Staff Payroll Ledger", columns)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_payroll_csv_bytes(payrolls):
    """
    Generates a CSV string buffer for staff payroll records.
    """
    output = io.StringIO()
    writer = csv.writer(output)

    columns = [
        "Payroll ID", "Staff ID", "Username", "Staff Name", "Designation",
        "Month/Year", "Basic Salary", "Allowances", "Bonus",
        "Deductions", "Net Salary", "Payment Status", "Payment Method", "Payment Date"
    ]
    writer.writerow(columns)

    for p in payrolls:
        staff = p.staff
        admin_user = staff.admin if staff else None
        designation = getattr(getattr(staff, 'salary_structure', None), 'designation', 'Staff')
        writer.writerow([
            p.id,
            staff.id if staff else "N/A",
            admin_user.username if admin_user else "N/A",
            f"{admin_user.first_name} {admin_user.last_name}".strip() if admin_user else "N/A",
            designation,
            f"{p.payroll_month}/{p.payroll_year}",
            f"{float(p.basic_salary or 0):.2f}",
            f"{float(p.allowances or 0):.2f}",
            f"{float(p.bonus or 0):.2f}",
            f"{float(p.deductions or 0):.2f}",
            f"{float(p.net_salary or 0):.2f}",
            p.payment_status,
            p.payment_method,
            str(p.payment_date) if p.payment_date else "N/A"
        ])

    return output.getvalue().encode('utf-8')



