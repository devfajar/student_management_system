import io
import openpyxl
from rest_framework import status
from django.urls import reverse
from student_management_app.tests.base import BaseAPITestCase
from student_management_app.models import (
    Attendance, AttendanceReport, FeeStructure, StudentFeeInvoice, StudentResult
)

class ReportsExcelPaginationAPITests(BaseAPITestCase):

    def setUp(self):
        super().setUp()
        # Seed test attendance
        self.attendance = Attendance.objects.create(
            subject_id=self.subject,
            attendance_date='2026-09-01',
            session_year_id=self.session_year
        )
        self.attendance_report = AttendanceReport.objects.create(
            student_id=self.student_profile,
            attendance_id=self.attendance,
            status=True
        )

        # Seed fee invoice
        self.fee_structure = FeeStructure.objects.create(
            course_id=self.course,
            session_year_id=self.session_year,
            fee_name='Annual Tuition 2026',
            tuition_fee=5000.0,
            due_date='2026-10-01'
        )
        self.invoice = StudentFeeInvoice.objects.create(
            student_id=self.student_profile,
            fee_structure_id=self.fee_structure,
            total_amount=5000.0,
            paid_amount=2000.0,
            payment_status='PARTIAL'
        )

        # Seed exam result
        self.result = StudentResult.objects.create(
            student_id=self.student_profile,
            subject_id=self.subject,
            subject_exam_marks=45.0,
            subject_assignment_marks=40.0
        )

    def test_admin_export_attendance_excel(self):
        self.authenticate_as_admin()
        url = reverse('export_attendance_excel')
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        self.assertIn('attachment; filename="attendance_report_', response['Content-Disposition'])

        # Validate Excel workbook structure
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        sheet = wb.active
        headers = [cell.value for cell in sheet[1]]
        self.assertIn("Student ID", headers)
        self.assertIn("Full Name", headers)
        self.assertIn("Subject", headers)
        self.assertIn("Status", headers)

    def test_admin_export_fees_excel(self):
        self.authenticate_as_admin()
        url = reverse('export_fees_excel')
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        sheet = wb.active
        headers = [cell.value for cell in sheet[1]]
        self.assertIn("Invoice ID", headers)
        self.assertIn("Student", headers)
        self.assertIn("Total Amount", headers)
        self.assertIn("Payment Status", headers)

    def test_admin_export_students_excel(self):
        self.authenticate_as_admin()
        url = reverse('export_students_excel')
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        sheet = wb.active
        headers = [cell.value for cell in sheet[1]]
        self.assertIn("Student ID", headers)
        self.assertIn("Username", headers)
        self.assertIn("Course", headers)

    def test_admin_export_results_excel(self):
        self.authenticate_as_admin()
        url = reverse('export_results_excel')
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        sheet = wb.active
        headers = [cell.value for cell in sheet[1]]
        self.assertIn("Student", headers)
        self.assertIn("Subject", headers)
        self.assertIn("Total Score", headers)

    def test_excel_export_filtered_by_search(self):
        self.authenticate_as_admin()
        # Search for student username
        url = f"{reverse('export_students_excel')}?search=test_student"
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        sheet = wb.active
        # Row 1 is header, Row 2 should be the student
        self.assertEqual(sheet.cell(row=2, column=2).value, 'test_student')

    def test_paginated_reports_preview_endpoint(self):
        self.authenticate_as_admin()
        url = reverse('reports_preview')
        response = self.client.get(f"{url}?type=students&page=1&page_size=10")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn('count', response.data)
        self.assertIn('total_pages', response.data)
        self.assertIn('current_page', response.data)
        self.assertIn('page_size', response.data)
        self.assertIn('results', response.data)
        self.assertTrue(len(response.data['results']) >= 1)

    def test_paginated_reports_search_filtering(self):
        self.authenticate_as_admin()
        url = reverse('reports_preview')
        response = self.client.get(f"{url}?type=fees&search=Annual Tuition")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(len(response.data['results']) >= 1)
        self.assertIn('Annual Tuition', response.data['results'][0]['fee_name'])

    def test_student_forbidden_from_admin_excel_exports(self):
        self.authenticate_as_student()
        url = reverse('export_attendance_excel')
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_unauthenticated_export_denied(self):
        url = reverse('export_attendance_excel')
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)
